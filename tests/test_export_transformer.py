"""Tests del exportador Excel y del sentimiento transformer (con fallback)."""

import sys
from pathlib import Path

RAIZ = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(RAIZ))

import exportar_excel
import sentimiento_politico as sp


def test_transformer_fallback_a_lexico(monkeypatch):
    # Forzar que el transformer no esté disponible → cae al léxico sin romper.
    monkeypatch.setattr(sp, "_cargar_transformer", lambda: None)
    r = sp.analizar_polaridad("Gran logro con apoyo y respaldo.", usar_transformer=True)
    assert r["polaridad"] == "positivo"  # vino del léxico


def test_lexico_por_defecto():
    r = sp.analizar_polaridad("Escándalo de corrupción y fraude grave.")
    assert r["polaridad"] == "negativo"


def test_exportar_excel(tmp_path):
    res = {
        "comparacion_candidatos": {
            "Cepeda": {
                "n_notas": 5,
                "polaridad": {"positivo": 3, "negativo": 1, "neutro": 1},
                "score_polaridad_medio": 0.3,
                "polarizacion_afectiva": 0.4,
                "encuadre_dominante": "Seguridad",
            }
        },
        "tendencia_medios": {
            "candidatos": ["Cepeda", "Espriella"],
            "medios": {
                "El Tiempo": {
                    "tono": {"Cepeda": 0.1, "Espriella": 0.3},
                    "sesgo": -0.2,
                    "favorece": "Espriella",
                }
            },
        },
        "por_nota": {
            "http://m/1": {
                "emociones": {
                    "polaridad": "positivo",
                    "score_polaridad": 0.3,
                    "emocion_dominante": "confianza",
                },
                "frame": {"etiqueta": "Seguridad"},
                "calidad": {"veredicto": "confiable", "score": 0.9},
            }
        },
        "series_temporales": {"meses": ["2026-06"], "volumen": {"2026-06": 1}},
        "frecuencias": [{"palabra": "voto", "freq": 10, "df": 5}],
        "topicos": {"topicos": [{"palabras": ["voto", "campaña"]}]},
        "cobertura_por_tipo": {"candidato": [{"actor": "Cepeda", "n_notas": 5}]},
        "calidad_corpus": {"confiable": 1, "revisar": 0, "malo": 0, "total": 1},
    }
    notas = [
        {
            "url": "http://m/1",
            "medio": "El Tiempo",
            "fecha_publicacion": "2026-06-10",
            "titular": "Nota",
        }
    ]
    ruta = exportar_excel.exportar(res, notas, tmp_path / "out.xlsx")
    assert ruta.exists()
    from openpyxl import load_workbook

    wb = load_workbook(ruta)
    # debe tener las hojas principales
    assert "Candidatos" in wb.sheetnames
    assert "Tendencia medios" in wb.sheetnames
    assert "Notas" in wb.sheetnames
