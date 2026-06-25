"""Exportación de resultados a Excel (.xlsx) para el paper.

Toma el dict de ``pipeline.analizar_corpus`` y genera un libro con una hoja por
dimensión: comparativa de candidatos, tendencia de medios, notas con su
clasificación, series temporales, frecuencias, tópicos y cobertura por tipo.
Listo para tablas y figuras de una publicación.
"""

from __future__ import annotations

from pathlib import Path


def _auto_ancho(ws):
    from openpyxl.utils import get_column_letter

    for col in ws.columns:
        ancho = max((len(str(c.value)) for c in col if c.value is not None), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(60, ancho + 2)


def exportar(res: dict, notas: list[dict], ruta: str | Path) -> Path:
    from openpyxl import Workbook
    from openpyxl.styles import Font

    ruta = Path(ruta)
    ruta.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    neg = Font(bold=True)

    def nueva(titulo, encabezados):
        ws = wb.create_sheet(titulo[:31])
        ws.append(encabezados)
        for c in ws[1]:
            c.font = neg
        return ws

    # quitar la hoja por defecto al final
    default = wb.active

    # 1) Comparativa de candidatos
    cand = res.get("comparacion_candidatos", {})
    if cand:
        ws = nueva(
            "Candidatos",
            [
                "Candidato",
                "Notas",
                "Positivo",
                "Negativo",
                "Neutro",
                "Tono medio",
                "Polariz. afectiva",
                "Encuadre dominante",
            ],
        )
        for n, d in sorted(cand.items(), key=lambda kv: -kv[1]["n_notas"]):
            p = d.get("polaridad", {})
            ws.append(
                [
                    n,
                    d["n_notas"],
                    p.get("positivo", 0),
                    p.get("negativo", 0),
                    p.get("neutro", 0),
                    d.get("score_polaridad_medio", 0),
                    d.get("polarizacion_afectiva", 0),
                    d.get("encuadre_dominante", ""),
                ]
            )
        _auto_ancho(ws)

    # 2) Tendencia / filiación de medios
    tm = res.get("tendencia_medios", {})
    if tm.get("medios"):
        cands = tm.get("candidatos", [])
        ws = nueva(
            "Tendencia medios", ["Medio"] + [f"Tono {c}" for c in cands] + ["Sesgo", "Favorece"]
        )
        for medio, d in tm["medios"].items():
            fila = (
                [medio]
                + [d["tono"].get(c) for c in cands]
                + [d.get("sesgo", 0), d.get("favorece") or ""]
            )
            ws.append(fila)
        _auto_ancho(ws)

    # 3) Notas con clasificación (datos crudos para reanálisis)
    por = res.get("por_nota", {})
    ws = nueva(
        "Notas",
        [
            "URL",
            "Medio",
            "Fecha",
            "Titular",
            "Polaridad",
            "Score",
            "Emoción",
            "Encuadre",
            "Calidad",
            "Score calidad",
        ],
    )
    for nt in notas:
        r = por.get(nt.get("url"), {})
        emo = r.get("emociones", {})
        cal = r.get("calidad", {})
        ws.append(
            [
                nt.get("url", ""),
                nt.get("medio", ""),
                nt.get("fecha_publicacion", ""),
                (nt.get("titular") or "")[:120],
                emo.get("polaridad", ""),
                emo.get("score_polaridad", ""),
                emo.get("emocion_dominante", ""),
                (r.get("frame", {}) or {}).get("etiqueta", ""),
                cal.get("veredicto", ""),
                cal.get("score", ""),
            ]
        )
    _auto_ancho(ws)

    # 4) Series temporales
    series = res.get("series_temporales", {})
    if series.get("meses"):
        ws = nueva("Series", ["Mes", "Volumen"])
        vol = series.get("volumen", {})
        for m in series["meses"]:
            ws.append([m, vol.get(m, 0)])
        _auto_ancho(ws)

    # 5) Frecuencias
    frec = res.get("frecuencias", [])
    if isinstance(frec, list) and frec:
        ws = nueva("Frecuencias", ["Término", "Frecuencia", "Docs"])
        for f in frec:
            ws.append([f.get("palabra", ""), f.get("freq", 0), f.get("df", "")])
        _auto_ancho(ws)

    # 6) Tópicos
    top = res.get("topicos", {})
    palabras_top = top.get("topicos") or top.get("topics")
    if isinstance(palabras_top, list):
        ws = nueva("Tópicos", ["Tópico", "Palabras clave"])
        for i, t in enumerate(palabras_top):
            pal = t.get("palabras") or t.get("words") or t
            ws.append(
                [f"Tópico {i + 1}", ", ".join(map(str, pal)) if isinstance(pal, list) else str(pal)]
            )
        _auto_ancho(ws)

    # 7) Cobertura por tipo (NER ricos)
    cob = res.get("cobertura_por_tipo", {})
    if cob:
        ws = nueva("Cobertura por tipo", ["Tipo", "Actor", "Nº notas"])
        for tipo, lst in cob.items():
            for x in lst:
                ws.append([tipo, x["actor"], x["n_notas"]])
        _auto_ancho(ws)

    # 8) Calidad del corpus (resumen)
    cq = res.get("calidad_corpus", {})
    if cq:
        ws = nueva("Calidad", ["Veredicto", "N"])
        for k in ("confiable", "revisar", "malo"):
            ws.append([k, cq.get(k, 0)])
        _auto_ancho(ws)

    wb.remove(default)
    wb.save(ruta)
    return ruta
